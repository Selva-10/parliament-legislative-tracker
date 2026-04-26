# tracker/views.py
from django.shortcuts import render, get_object_or_404
from django.views.generic import TemplateView
from django.core.paginator import Paginator
from django.db import models
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import JsonResponse
from django.db.models import Count

import json
import csv
import io

# For Excel download
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# For PDF download
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from .models import Bill, State, Party, MP, MLA

# -------------------------------------------------------------------
# Dashboard
# -------------------------------------------------------------------
class DashboardView(TemplateView):
    template_name = 'tracker/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Bills
        context['total_bills'] = Bill.objects.count()
        context['pending_bills'] = Bill.objects.filter(status='PENDING').count()
        context['passed_bills'] = Bill.objects.filter(status='PASSED').count()
        context['rejected_bills'] = Bill.objects.filter(status='REJECTED').count()
        context['loksabha_bills'] = Bill.objects.filter(house='LOK_SABHA').count()
        context['rajyasabha_bills'] = Bill.objects.filter(house='RAJYA_SABHA').count()
        context['recent_bills'] = Bill.objects.order_by('-introduction_date')[:10]

        # MPs & MLAs
        context['total_mps'] = MP.objects.count()
        context['total_mlas'] = MLA.objects.count()
        context['recent_mps'] = MP.objects.select_related('state', 'party').order_by('-id')[:5]
        context['recent_mlas'] = MLA.objects.select_related('state', 'party').order_by('-id')[:5]
        return context

# -------------------------------------------------------------------
# Bills
# -------------------------------------------------------------------
def bill_list(request):
    bills = Bill.objects.all().order_by('-introduction_date')

    search_query = request.GET.get('search', '')
    if search_query:
        bills = bills.filter(
            models.Q(title__icontains=search_query) |
            models.Q(bill_id__icontains=search_query) |
            models.Q(bill_number__icontains=search_query) |
            models.Q(introduced_by__icontains=search_query)
        )

    state_param = request.GET.get('state', '')
    if state_param:
        bills = bills.filter(state__icontains=state_param)

    house_param = request.GET.get('house', '')
    house_map = {'Lok Sabha': 'LOK_SABHA', 'Rajya Sabha': 'RAJYA_SABHA', 'Both': 'BOTH'}
    reverse_map = {v: k for k, v in house_map.items()}
    if house_param:
        db_house = house_map.get(house_param, house_param)
        bills = bills.filter(house=db_house)

    status_param = request.GET.get('status', '')
    if status_param:
        bills = bills.filter(status=status_param.upper())

    paginator = Paginator(bills, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    db_houses = set(Bill.objects.exclude(house__isnull=True).exclude(house='')
                    .values_list('house', flat=True).distinct())
    valid_houses = {'LOK_SABHA', 'RAJYA_SABHA', 'BOTH'}
    houses_to_show = [reverse_map.get(h, h.title().replace('_', ' ')) for h in db_houses if h in valid_houses]

    db_statuses = set(Bill.objects.exclude(status__isnull=True).exclude(status='')
                      .values_list('status', flat=True).distinct())
    valid_statuses = {'PENDING', 'PASSED', 'REJECTED', 'WITHDRAWN', 'LAPSED'}
    statuses_to_show = [s for s in db_statuses if s in valid_statuses]

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'selected_house': house_param,
        'selected_status': status_param,
        'selected_state': state_param,
        'houses': houses_to_show,
        'statuses': statuses_to_show,
        'total_count': bills.count(),
    }
    return render(request, 'tracker/bill_list.html', context)


def bill_detail(request, pk):
    bill = get_object_or_404(Bill, pk=pk)
    return render(request, 'tracker/bill_detail.html', {'bill': bill})

# -------------------------------------------------------------------
# Analytics
# -------------------------------------------------------------------
def analytics(request):
    from django.db.models import Count, Q
    from datetime import date, timedelta
    import json
    
    # Get year filter from request
    selected_year = request.GET.get('year', 'all')
    
    # Base queryset
    mpa_bills = Bill.objects.filter(source='MPA')
    
    # Apply year filter
    if selected_year != 'all' and selected_year:
        mpa_bills = mpa_bills.filter(introduction_date__year=int(selected_year))
    
    # Statistics Cards
    total_bills = mpa_bills.count()
    ls_count = mpa_bills.filter(house='LOK_SABHA').count()
    rs_count = mpa_bills.filter(house='RAJYA_SABHA').count()
    both_count = mpa_bills.filter(house='BOTH').count()
    
    # Bills by Status
    status_counts = list(mpa_bills.values('status').annotate(count=Count('status')))
    
    # Bills by House
    house_counts = [
        {'house': 'Lok Sabha', 'count': ls_count},
        {'house': 'Rajya Sabha', 'count': rs_count},
        {'house': 'Both', 'count': both_count},
    ]
    
    # Top Ministries
    ministry_counts = list(mpa_bills.exclude(ministry__isnull=True).exclude(ministry='')
                           .values('ministry').annotate(count=Count('ministry')).order_by('-count')[:10])
    
    # ========== MONTHLY DATA ==========
    monthly_data = []
    first_bill = Bill.objects.filter(source='MPA').order_by('introduction_date').first()
    last_bill = Bill.objects.filter(source='MPA').order_by('-introduction_date').first()
    
    if first_bill and last_bill and first_bill.introduction_date and last_bill.introduction_date:
        start_year = first_bill.introduction_date.year
        end_year = last_bill.introduction_date.year
        
        if selected_year != 'all' and selected_year:
            start_year = int(selected_year)
            end_year = int(selected_year)
        
        for year in range(start_year, end_year + 1):
            for month in range(1, 13):
                month_start = date(year, month, 1)
                
                if month == 12:
                    month_end = date(year + 1, 1, 1) - timedelta(days=1)
                else:
                    month_end = date(year, month + 1, 1) - timedelta(days=1)
                
                month_bills = mpa_bills.filter(
                    introduction_date__gte=month_start,
                    introduction_date__lte=month_end
                )
                
                bill_count = month_bills.count()
                
                if bill_count > 0:
                    monthly_data.append({
                        'month': month_start.strftime('%b %Y'),
                        'total': bill_count,
                        'ls_count': month_bills.filter(house='LOK_SABHA').count(),
                        'rs_count': month_bills.filter(house='RAJYA_SABHA').count(),
                        'passed': month_bills.filter(status='PASSED').count(),
                        'pending': month_bills.filter(status='PENDING').count(),
                        'rejected': month_bills.filter(status__in=['REJECTED', 'WITHDRAWN', 'LAPSED']).count(),
                    })
    
    # ========== MINISTRY SUCCESS RATE WITH YEAR-WISE BREAKDOWN ==========
    ministry_success = []
    
    # Get available years for each ministry
    all_years = Bill.objects.filter(source='MPA', introduction_date__isnull=False) \
                    .dates('introduction_date', 'year') \
                    .values_list('introduction_date__year', flat=True).distinct().order_by('introduction_date__year')
    years_list = sorted(set(all_years))
    
    for mc in ministry_counts:
        ministry = mc['ministry']
        total = mc['count']
        passed = mpa_bills.filter(ministry=ministry, status='PASSED').count()
        success_rate = (passed / total * 100) if total > 0 else 0
        
        # Get year-wise breakdown for this ministry
        year_wise = []
        for year in years_list:
            year_bills = Bill.objects.filter(source='MPA', ministry=ministry, introduction_date__year=year)
            year_total = year_bills.count()
            year_passed = year_bills.filter(status='PASSED').count()
            if year_total > 0:
                year_wise.append({
                    'year': year,
                    'total': year_total,
                    'passed': year_passed,
                    'success_rate': round((year_passed / year_total * 100), 1)
                })
        
        ministry_success.append({
            'ministry': ministry,
            'total': total,
            'passed': passed,
            'success_rate': round(success_rate, 1),
            'year_wise': year_wise,
        })
    
    # Get available years for filter dropdown
    available_years = Bill.objects.filter(source='MPA', introduction_date__isnull=False) \
                         .dates('introduction_date', 'year') \
                         .values_list('introduction_date__year', flat=True).distinct().order_by('-introduction_date__year')
    years_filter = sorted(set(available_years), reverse=True)
    
    context = {
        'total_bills': total_bills,
        'ls_count': ls_count,
        'rs_count': rs_count,
        'both_count': both_count,
        'status_counts': json.dumps(status_counts),
        'house_counts': json.dumps(house_counts),
        'ministry_counts': json.dumps(ministry_counts),
        'ministry_success': json.dumps(ministry_success),
        'monthly_data': json.dumps(monthly_data),
        'years': years_filter,
        'selected_year': selected_year,
        'all_years': list(years_list),
    }
    return render(request, 'tracker/analytics.html', context)

# -------------------------------------------------------------------
# Map
# -------------------------------------------------------------------
def map_view(request):
    from django.db.models import Count

    status_filter = request.GET.get('status', 'all')
    year_filter = request.GET.get('year', 'all')

    bills = Bill.objects.all()
    if status_filter != 'all':
        bills = bills.filter(status=status_filter)
    if year_filter != 'all' and year_filter:
        bills = bills.filter(introduction_date__year=year_filter)

    state_data = {}
    for bill in bills:
        state = bill.state or 'Other'
        if state not in state_data:
            state_data[state] = {'count': 0, 'parties': {}}
        state_data[state]['count'] += 1
        party = bill.introduced_by_party or 'Unknown'
        state_data[state]['parties'][party] = state_data[state]['parties'].get(party, 0) + 1

    states_list = []
    for state, data in state_data.items():
        dominant_party = max(data['parties'].items(), key=lambda x: x[1])[0] if data['parties'] else 'Unknown'
        states_list.append({'state': state, 'count': data['count'], 'party': dominant_party})

    years = Bill.objects.exclude(introduction_date__isnull=True).dates('introduction_date', 'year')
    years_list = [y.year for y in years]

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'states': states_list, 'total_bills': bills.count()})

    state_coords = {
        'Delhi': [28.6139, 77.2090],
        'Maharashtra': [19.7515, 75.7139],
        'Uttar Pradesh': [26.8467, 80.9462],
        'West Bengal': [22.9868, 87.8550],
        'Karnataka': [15.3173, 75.7139],
        'Tamil Nadu': [11.1271, 78.6569],
        'Gujarat': [22.2587, 71.1924],
        'Bihar': [25.0961, 85.3131],
        'Madhya Pradesh': [23.2599, 77.4126],
        'Rajasthan': [27.0238, 74.2179],
        'Punjab': [31.1471, 75.3412],
        'Odisha': [20.9517, 85.0985],
        'Jharkhand': [23.6102, 85.2799],
        'Assam': [26.2006, 92.9376],
        'Telangana': [17.1232, 79.2088],
        'Other': [22.0, 79.0],
    }

    context = {
        'states_data': json.dumps(states_list),
        'state_coords': json.dumps(state_coords),
        'total_bills': bills.count(),
        'status_filter': status_filter,
        'year_filter': year_filter,
        'years': years_list,
    }
    return render(request, 'tracker/map.html', context)

# -------------------------------------------------------------------
# Calendar
# -------------------------------------------------------------------
def calendar_view(request):
    year = request.GET.get('year')
    month = request.GET.get('month')
    bills = Bill.objects.exclude(introduction_date__isnull=True)
    if year and month:
        bills = bills.filter(introduction_date__year=year, introduction_date__month=month)

    events = []
    for bill in bills:
        if bill.introduction_date:
            events.append({
                'title': f"{bill.bill_id}: {bill.title[:30]}",
                'start': bill.introduction_date.isoformat(),
                'url': f"/bills/{bill.id}/",
                'color': '#28a745' if bill.status == 'PASSED' else '#ffc107' if bill.status == 'PENDING' else '#dc3545'
            })

    context = {
        'events': json.dumps(events),
        'year': year,
        'month': month,
        'total_bills': Bill.objects.count(),
        'passed_count': Bill.objects.filter(status='PASSED').count(),
        'pending_count': Bill.objects.filter(status='PENDING').count(),
        'rejected_count': Bill.objects.filter(status='REJECTED').count() + Bill.objects.filter(status='WITHDRAWN').count(),
    }
    return render(request, 'tracker/calendar.html', context)

# -------------------------------------------------------------------
# API
# -------------------------------------------------------------------
def api_bills(request):
    from django.utils import timezone

    bills = Bill.objects.all().order_by('-introduction_date')
    house = request.GET.get('house')
    if house and house != 'all':
        bills = bills.filter(house=house)
    status = request.GET.get('status')
    if status and status != 'all':
        bills = bills.filter(status=status)
    search = request.GET.get('search')
    if search:
        bills = bills.filter(
            models.Q(title__icontains=search) |
            models.Q(bill_id__icontains=search)
        )
    date = request.GET.get('date')
    if date:
        bills = bills.filter(introduction_date=date)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        bills = bills.filter(introduction_date__gte=start_date, introduction_date__lte=end_date)
    limit = int(request.GET.get('limit', 500))
    bills = bills[:limit]

    today = timezone.now().date()
    data = [{
        'id': str(bill.id),
        'bill_id': bill.bill_id,
        'bill_number': bill.bill_number or '',
        'title': bill.title,
        'house': bill.house,
        'status': bill.status,
        'introduction_date': bill.introduction_date.isoformat() if bill.introduction_date else None,
        'introduced_by': bill.introduced_by or '',
        'ministry': bill.ministry or '',
        'url': f"/bills/{bill.id}/",
        'is_upcoming': bill.introduction_date > today if bill.introduction_date else False,
    } for bill in bills]
    return JsonResponse({'bills': data, 'count': len(data), 'total': Bill.objects.count()})

# -------------------------------------------------------------------
# download
# -------------------------------------------------------------------
def download_page(request):
    ministries = Bill.objects.exclude(ministry__isnull=True).exclude(ministry='')\
                     .values_list('ministry', flat=True).distinct().order_by('ministry')
    parties = Bill.objects.exclude(introduced_by_party__isnull=True).exclude(introduced_by_party='')\
                  .values_list('introduced_by_party', flat=True).distinct().order_by('introduced_by_party')
    years = Bill.objects.exclude(introduction_date__isnull=True).dates('introduction_date', 'year')
    years_list = [y.year for y in years]
    return render(request, 'tracker/download.html', {
        'ministries': ministries,
        'parties': parties,
        'years': years_list,
    })


def download_bills(request):
    from django.db import models

    bills = Bill.objects.all().order_by('-introduction_date')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    ministry = request.GET.get('ministry')
    party = request.GET.get('party')
    house = request.GET.get('house')
    status = request.GET.get('status')

    if start_date:
        bills = bills.filter(introduction_date__gte=start_date)
    if end_date:
        bills = bills.filter(introduction_date__lte=end_date)
    if ministry:
        bills = bills.filter(ministry=ministry)
    if party:
        bills = bills.filter(introduced_by_party=party)
    if house and house != 'all':
        bills = bills.filter(house=house)
    if status and status != 'all':
        bills = bills.filter(status=status)

    data = []
    for bill in bills:
        data.append({
            'bill_id': bill.bill_id,
            'bill_number': bill.bill_number or '',
            'title': bill.title,
            'house': bill.get_house_display() if hasattr(bill, 'get_house_display') else bill.house,
            'status': bill.get_status_display() if hasattr(bill, 'get_status_display') else bill.status,
            'introduced_by': bill.introduced_by or '',
            'introduced_by_party': bill.introduced_by_party or '',
            'introduction_date': bill.introduction_date.strftime('%d %b %Y') if bill.introduction_date else '',
            'ministry': bill.ministry or '',
            'state': bill.state or '',
            'source': bill.source or '',
        })

    fmt = request.GET.get('format', 'csv')
    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="bills.csv"'
        if data:
            writer = csv.DictWriter(response, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        return response
    elif fmt == 'json':
        return JsonResponse(data, safe=False)
    elif fmt == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Bills"
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            for row, bill in enumerate(data, 2):
                for col, key in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=bill[key])
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="bills.xlsx"'
        wb.save(response)
        return response
    elif fmt == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        styles = getSampleStyleSheet()
        title = Paragraph("Bills Download", styles['Title'])
        elements.append(title)
        if data:
            headers = list(data[0].keys())
            table_data = [headers] + [[bill[h] for h in headers] for bill in data]
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="bills.pdf"'
        response.write(pdf)
        return response
    else:
        return HttpResponse("Invalid format", status=400)

# -------------------------------------------------------------------
# Scrape trigger (if you have it)
# -------------------------------------------------------------------
def trigger_scrape(request):
    if request.method == 'POST':
        source = request.POST.get('source', 'all')
        try:
            from .scraper import RealBillScraper
            scraper = RealBillScraper()
            if source == 'all':
                result = scraper.scrape_all()
                message = "Scraping completed for all sources"
            elif source == 'MPA':
                result = scraper.scrape_mpa_bills()
                message = f"MPA scraping completed: {len(result)} bills"
            else:
                return JsonResponse({'status': 'error', 'message': 'Invalid source'})
            return JsonResponse({'status': 'success', 'message': message, 'result': result})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Scraping failed: {str(e)}'})
    sources = [
        {'id': 'all', 'name': 'All Sources'},
        {'id': 'MPA', 'name': 'MPA (Ministry of Parliamentary Affairs)'},
    ]
    return render(request, 'tracker/scrape.html', {'sources': sources})

# -------------------------------------------------------------------
# MPs and MLAs
# -------------------------------------------------------------------
def mp_list(request):
    mps = MP.objects.select_related('state', 'party').all()
    state = request.GET.get('state')
    party = request.GET.get('party')
    house = request.GET.get('house')
    if state:
        mps = mps.filter(state__name__icontains=state)
    if party:
        mps = mps.filter(party__name__icontains=party)
    if house:
        mps = mps.filter(house=house)
    states = State.objects.filter(mps__isnull=False).distinct()
    parties = Party.objects.filter(mps__isnull=False).distinct()
    return render(request, 'tracker/mp_list.html', {
        'mps': mps,
        'states': states,
        'parties': parties,
    })

def mp_detail(request, pk):
    mp = get_object_or_404(MP, pk=pk)
    return render(request, 'tracker/mp_detail.html', {'mp': mp})

def mla_list(request):
    mlas = MLA.objects.select_related('state', 'party').all()
    state = request.GET.get('state')
    party = request.GET.get('party')
    if state:
        mlas = mlas.filter(state__name__icontains=state)
    if party:
        mlas = mlas.filter(party__name__icontains=party)
    states = State.objects.filter(mlas__isnull=False).distinct()
    parties = Party.objects.filter(mlas__isnull=False).distinct()
    return render(request, 'tracker/mla_list.html', {
        'mlas': mlas,
        'states': states,
        'parties': parties,
    })

def mla_detail(request, pk):
    mla = get_object_or_404(MLA, pk=pk)
    return render(request, 'tracker/mla_detail.html', {'mla': mla})
# Add to your bill_list view in views.py
from django.db import models

# tracker/views.py - Corrected bill_list function


# tracker/views.py - Modified bill_list for ONLY MPA bills

def bill_list(request):
    # Show ONLY MPA bills (central bills, not state bills)
    bills = Bill.objects.filter(source='MPA').order_by('-introduction_date')
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    selected_house = request.GET.get('house', '')
    selected_status = request.GET.get('status', '')
    
    # Apply filters
    if search_query:
        bills = bills.filter(
            models.Q(title__icontains=search_query) |
            models.Q(bill_id__icontains=search_query) |
            models.Q(bill_number__icontains=search_query)
        )
    
    if selected_house:
        bills = bills.filter(house=selected_house)
    
    if selected_status:
        bills = bills.filter(status=selected_status.upper())
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(bills, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get filter options (only from MPA bills)
    houses = Bill.objects.filter(source='MPA').values_list('house', flat=True).distinct()
    statuses = Bill.objects.filter(source='MPA').values_list('status', flat=True).distinct()
    
    context = {
        'page_obj': page_obj,
        'total_count': bills.count(),
        'search_query': search_query,
        'selected_house': selected_house,
        'selected_status': selected_status,
        'houses': houses,
        'statuses': statuses,
    }
    return render(request, 'tracker/bill_list.html', context)
    


# tracker/views.py - Map API for ONLY state bills
from django.http import JsonResponse
from django.db.models import Count
from tracker.models import Bill
import uuid

def api_state_bill_counts(request):
    """API to get bill counts per state"""
    from django.db.models import Count
    
    state_counts = Bill.objects.filter(
        source='STATE_BILL',
        state__isnull=False
    ).exclude(state='').values('state').annotate(count=Count('id'))
    
    result = [{'name': item['state'], 'count': item['count']} for item in state_counts]
    return JsonResponse(result, safe=False)


def api_state_bills(request, state_name):
    """API to get bills for a specific state"""
    bills = Bill.objects.filter(
        source='STATE_BILL',
        state__icontains=state_name
    ).values('id', 'title', 'introduction_date', 'bill_id', 'legislative_year')
    
    return JsonResponse({
        'state': state_name,
        'count': bills.count(),
        'bills': list(bills)
    })


def api_bill_detail(request, bill_id):
    """API to get bill details by bill_id (e.g., KAR-0001)"""
    from django.http import JsonResponse
    from tracker.models import Bill
    
    try:
        # Get bill by bill_id field (KAR-0001, KAR-0002, etc.)
        bill = Bill.objects.get(bill_id=bill_id)
        
        return JsonResponse({
            'id': str(bill.id),
            'bill_id': bill.bill_id,
            'title': bill.title,
            'state': bill.state if bill.state else 'Karnataka',
            'introduction_date': bill.introduction_date.strftime('%Y-%m-%d') if bill.introduction_date else None,
            'status': bill.status,
            'source': bill.source,
        })
        
    except Bill.DoesNotExist:
        return JsonResponse({'error': f'Bill not found: {bill_id}'}, status=404)
    
def india_map(request):
    """Render India map page"""
    return render(request, 'tracker/india_map.html')

def state_bills_list(request, state_name):
    """Display state bills with simplified columns"""
    bills = Bill.objects.filter(
        source='STATE_BILL',
        state__icontains=state_name
    ).order_by('-legislative_year')
    
    context = {
        'bills': bills,
        'state_name': state_name,
        'total_count': bills.count(),
    }
    return render(request, 'tracker/state_bills_list.html', context)
def api_bill_detail_by_bill_id(request, bill_id):
    """API to get bill details by bill_id (e.g., KAR-0001)"""
    from django.http import JsonResponse
    from tracker.models import Bill
    
    print(f"🔍 Looking for bill with bill_id: {bill_id}")
    
    try:
        bill = Bill.objects.get(bill_id=bill_id)
        
        return JsonResponse({
            'id': str(bill.id),
            'bill_id': bill.bill_id,
            'title': bill.title,
            'state': bill.state,
            'introduction_date': bill.introduction_date.strftime('%Y-%m-%d') if bill.introduction_date else None,
            'status': bill.status,
            'source': bill.source,
        })
        
    except Bill.DoesNotExist:
        return JsonResponse({'error': 'Bill not found'}, status=404)